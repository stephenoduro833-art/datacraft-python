from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sales Report"

ws["A1"] = "Business"
ws["B1"] = "Revenue"
ws["C1"] = "Expenses"
ws["D1"] = "Profit"

ws.append(["Datacraft", 5000, 3200, 1800])
ws.append(["Lunava", 8000, 5500, 2500])
ws.append(["TechHub", 12000, 7500, 4500])
from openpyxl.styles import Font, PatternFill

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="1a1a2e")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
from openpyxl.styles import PatternFill

light_fill = PatternFill(fill_type="solid", fgColor="f4f4f4")
white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
    fill = light_fill if i % 2 == 0 else white_fill
    for cell in row:
        cell.fill = fill
wb.save("openpyxl_report.xlsx")
print("File created")