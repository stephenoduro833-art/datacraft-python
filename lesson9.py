import mysql.connector
import pandas as pd

connection = mysql.connector.connect(
    host="localhost",
    user="root",
    password="datacraft",
    database="datacraft_practice"
)

print("Connected successfully")

cursor = connection.cursor()
cursor.execute("SELECT * FROM customers")
results = cursor.fetchall()

df = pd.DataFrame(results, columns=["ID", "Name", "Phone", "City"])
print(df)
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

df.to_excel("customers_report.xlsx", index=False)

wb = load_workbook("customers_report.xlsx")
ws = wb.active

# Bold headers with dark background
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="1a1a2e")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Column widths
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15

wb.save("customers_report.xlsx")
print("Report exported")