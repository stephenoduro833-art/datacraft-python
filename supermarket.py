import mysql.connector
import pandas as pd

connection = mysql.connector.connect(
    host="localhost",
    user="root",
    password="datacraft",
    database="supermarket_db"
)

query = """
SELECT 
    s.sale_id,
    s.sale_date,
    s.sale_time,
    c.customer_name,
    ca.cashier_name,
    b.branch_name,
    p.product_name,
    p.category,
    si.quantity,
    si.unit_price,
    (si.quantity * si.unit_price) AS total
FROM sales s
JOIN customers c ON s.customer_id = c.customer_id
JOIN cashiers ca ON s.cashier_id = ca.cashier_id
JOIN branches b ON s.branch_id = b.branch_id
JOIN sale_items si ON s.sale_id = si.sale_id
JOIN products p ON si.product_id = p.product_id
ORDER BY s.sale_date, s.sale_time
"""

df = pd.read_sql(query, connection)
print(df)
branch_revenue = df.groupby("branch_name")["total"].sum().reset_index()
branch_revenue.columns = ["Branch", "Total Revenue"]
branch_revenue = branch_revenue.sort_values("Total Revenue", ascending=False)
print("\n--- Revenue by Branch ---")
print(branch_revenue)

product_sales = df.groupby("product_name")["quantity"].sum().reset_index()
product_sales.columns = ["Product", "Total Quantity Sold"]
product_sales = product_sales.sort_values("Total Quantity Sold", ascending=False)
print("\n--- Top Selling Products ---")
print(product_sales)

category_revenue = df.groupby("category")["total"].sum().reset_index()
category_revenue.columns = ["Category", "Total Revenue"]
category_revenue = category_revenue.sort_values("Total Revenue", ascending=False)
print("\n--- Revenue by Category ---")
print(category_revenue)

cashier_performance = df.groupby("cashier_name")["sale_id"].nunique().reset_index()
cashier_performance.columns = ["Cashier", "Total Transactions"]
cashier_performance = cashier_performance.sort_values("Total Transactions", ascending=False)
print("\n--- Cashier Performance ---")
print(cashier_performance)


df["hour"] = df["sale_time"].apply(lambda x: x.seconds // 3600)
hourly_traffic = df.groupby("hour")["sale_id"].nunique().reset_index()
hourly_traffic.columns = ["Hour", "Transactions"]
hourly_traffic = hourly_traffic.sort_values("Transactions", ascending=False)
print("\n--- Busiest Hours ---")
print(hourly_traffic)

inventory_query = """
SELECT 
    p.product_name,
    p.category,
    i.quantity AS current_stock,
    p.reorder_level,
    b.branch_name
FROM inventory i
JOIN products p ON i.product_id = p.product_id
JOIN branches b ON i.branch_id = b.branch_id
WHERE i.quantity <= p.reorder_level
ORDER BY i.quantity ASC
"""

reorder_df = pd.read_sql(inventory_query, connection)
print("\n--- Products Needing Reorder ---")
print(reorder_df)

df["sale_time"] = df["sale_time"].apply(lambda x: str(x).split(" ")[-1][:5])
hourly_traffic["Hour"] = hourly_traffic["Hour"].apply(lambda x: f"{x:02d}:00 AM" if x < 12 else f"{x-12:02d}:00 PM" if x > 12 else "12:00 PM")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

with pd.ExcelWriter("supermarket_report.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Master Data", index=False)
    branch_revenue.to_excel(writer, sheet_name="Branch Revenue", index=False)
    product_sales.to_excel(writer, sheet_name="Product Sales", index=False)
    category_revenue.to_excel(writer, sheet_name="Category Revenue", index=False)
    cashier_performance.to_excel(writer, sheet_name="Cashier Performance", index=False)
    hourly_traffic.to_excel(writer, sheet_name="Hourly Traffic", index=False)
    reorder_df.to_excel(writer, sheet_name="Reorder Alert", index=False)

wb = load_workbook("supermarket_report.xlsx")

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(fill_type="solid", fgColor="1a1a2e")
alt_fill = PatternFill(fill_type="solid", fgColor="f4f4f4")

for ws in wb.worksheets:
    for cell in ws[1]:
        ws.freeze_panes = "A2"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        for cell in row:
            if i % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

wb.save("supermarket_report.xlsx")
print("Report formatted and saved")

